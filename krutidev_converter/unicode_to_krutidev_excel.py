#!/usr/bin/env python3
"""
Unicode (Devanagari) → KrutiDev 10 Converter
=============================================
Fully self-contained — no external files, no internet, no git clone needed.
Reads an Excel file, converts all Hindi/Devanagari text columns to KrutiDev 10
encoding, and saves a new Excel file with _KD suffix columns added.

Build to exe:
    pip install pyinstaller
    pyinstaller --onefile --windowed krutidev_converter.py

Requirements:
    pip install openpyxl pandas
"""

import sys, os, threading
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# Unicode → KrutiDev conversion tables
# ─────────────────────────────────────────────────────────────────────────────

# Conjuncts must be checked BEFORE single-character mappings
CONJUNCTS = {
    "क्ष": "f{k",   "त्र": "\x8c",  "ज्ञ": "\xab",  "श्र": ";z",
    "स्त": "Lr",    "स्थ": "LFk",   "प्र": "iz",    "क्र": "\xd8z",
    "ग्र": "xz",    "द्र": "nz",    "ब्र": "cz",    "भ्र": "Hkz",
    "ह्म": "\xe1",  "ह्न": "gu",    "द्ध": "f)",    "त्त": "\xd9k",
    "न्न": "Uu",    "स्न": "Lu",    "क्त": "Dr",    "ग्न": "Xu",
}

U2K = {
    # Independent vowels
    "अ": "v",   "आ": "vk",  "इ": "b",   "ई": "bZ",  "उ": "m",
    "ऊ": "Å",   "ऋ": "_",   "ए": ",",   "ऐ": ",s",  "ओ": "vks",
    "औ": "vkS",
    # Consonants
    "क": "d",   "ख": "[k",  "ग": "x",   "घ": "?k",  "ङ": "M~",
    "च": "p",   "छ": "N",   "ज": "t",   "झ": ">",   "ञ": "¥",
    "ट": "V",   "ठ": "B",   "ड": "M",   "ढ": "<",   "ण": ".k",
    "त": "r",   "थ": "Fk",  "द": "n",   "ध": "/k",  "न": "u",
    "प": "i",   "फ": "Q",   "ब": "c",   "भ": "Hk",  "म": "e",
    "य": "';",  "र": "j",   "ल": "y",   "व": "o",   "श": "'\\k",
    "ष": "\"k", "स": "l",   "ह": "g",
    # Nukta variants
    "ज़": "t+",  "ड़": "M+",  "ढ़": "<+",  "फ़": "Q+",
    # Matras (vowel signs)
    "ा": "k",   "ि": "f",   "ी": "h",   "ु": "q",   "ू": "w",
    "ृ": "^",   "े": "s",   "ै": "S",   "ो": "ks",  "ौ": "kS",
    "ं": "a",   "ः": "%",   "ँ": "W",   "़": "+",   "्": "~",
    # Punctuation
    "।": ",",   "॥": ",,",  "ॐ": "ÅA",
    # Devanagari digits → ASCII
    "०": "0",   "१": "1",   "२": "2",   "३": "3",   "४": "4",
    "५": "5",   "६": "6",   "७": "7",   "८": "8",   "९": "9",
    # Zero-width joiners (drop)
    "\u200c": "", "\u200d": "",
}


def unicode_to_krutidev(text: str) -> str:
    """Convert a single Unicode/Devanagari string to KrutiDev 10 encoding."""
    if not isinstance(text, str) or not text.strip():
        return text
    tokens, i, chars = [], 0, list(text)
    while i < len(chars):
        matched = False
        for length in (3, 2):
            if i + length <= len(chars):
                seg = "".join(chars[i : i + length])
                if seg in CONJUNCTS:
                    tokens.append(CONJUNCTS[seg])
                    i += length
                    matched = True
                    break
        if not matched:
            tokens.append(U2K.get(chars[i], chars[i]))
            i += 1
    return _fix_i_matra("".join(tokens))


def _fix_i_matra(text: str) -> str:
    """Swap 'f' (ि matra) one position left — KrutiDev rendering convention."""
    out = list(text)
    i = 1
    while i < len(out):
        if out[i] == "f":
            out[i], out[i - 1] = out[i - 1], out[i]
        i += 1
    return "".join(out)


# ─────────────────────────────────────────────────────────────────────────────
# Excel processing
# ─────────────────────────────────────────────────────────────────────────────

def process_excel(input_path: str, output_path: str,
                  columns: list | None = None,
                  progress_cb=None) -> tuple[int, int]:
    """
    Read input Excel, convert chosen columns, write output Excel.
    Returns (rows_processed, columns_converted).
    progress_cb(pct: int, msg: str) is called during processing if provided.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    def _prog(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    _prog(5, "Reading Excel…")
    wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws_in = wb_in.active
    rows_raw = list(ws_in.iter_rows(values_only=True))
    wb_in.close()

    if not rows_raw:
        raise ValueError("Excel file is empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows_raw[0]]
    data    = rows_raw[1:]

    # Decide which columns to convert
    if columns:
        to_convert = [c for c in columns if c in headers]
    else:
        # Auto: all columns that contain any Devanagari text
        to_convert = []
        for hi, hdr in enumerate(headers):
            for row in data[:50]:        # sample first 50 rows
                val = row[hi] if hi < len(row) else None
                if val and isinstance(val, str):
                    if any("\u0900" <= ch <= "\u097f" for ch in val):
                        to_convert.append(hdr)
                        break

    if not to_convert:
        raise ValueError(
            "No Devanagari columns found.\n"
            "Make sure the Excel file contains Hindi (Unicode) text.")

    _prog(15, f"Found {len(to_convert)} Hindi column(s) — converting…")

    # Build new column list: original + _KD after each converted column
    new_headers = []
    kd_positions = {}   # original col name → index of its _KD col in new_headers
    for hdr in headers:
        new_headers.append(hdr)
        if hdr in to_convert:
            kd_name = f"{hdr}_KD"
            kd_positions[hdr] = len(new_headers)
            new_headers.append(kd_name)

    # Convert data
    new_rows = []
    total = len(data)
    for ri, row in enumerate(data):
        new_row = []
        for hi, hdr in enumerate(headers):
            val = row[hi] if hi < len(row) else None
            new_row.append(val)
            if hdr in to_convert:
                kd_val = unicode_to_krutidev(str(val)) if val is not None else ""
                new_row.append(kd_val)
        new_rows.append(new_row)
        if ri % 200 == 0:
            pct = 15 + int(70 * ri / max(total, 1))
            _prog(pct, f"Converting row {ri+1:,} / {total:,}…")

    _prog(85, "Writing output Excel…")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    # Write and format headers
    blue  = PatternFill("solid", fgColor="4472C4")
    green = PatternFill("solid", fgColor="C6EFCE")
    ws_out.append(new_headers)
    for cell in ws_out[1]:
        name = cell.value or ""
        is_kd = name.endswith("_KD")
        cell.fill  = green if is_kd else blue
        cell.font  = Font(
            bold=True,
            color="000000" if is_kd else "FFFFFF",
            name="Kruti Dev 010" if is_kd else "Calibri")
        cell.alignment = Alignment(horizontal="center")
        ws_out.column_dimensions[cell.column_letter].width = max(18, len(name) + 4)

    # Write data rows
    for new_row in new_rows:
        ws_out.append(new_row)

    # Apply Kruti Dev 010 font to all _KD data cells
    kd_col_indices = {i+1 for i, h in enumerate(new_headers) if h.endswith("_KD")}
    for ri, row_cells in enumerate(
            ws_out.iter_rows(min_row=2, max_row=len(new_rows)+1), start=2):
        for cell in row_cells:
            if cell.column in kd_col_indices:
                cell.font = Font(name="Kruti Dev 010", size=11)

    wb_out.save(output_path)
    _prog(100, f"Done — {total:,} rows, {len(to_convert)} column(s) converted.")
    return total, len(to_convert)


# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────

try:
    import tkinter as tk
    from tkinter import ttk, messagebox
    from tkinter.filedialog import askopenfilename, asksaveasfilename
except ImportError:
    print("ERROR: tkinter not available.")
    sys.exit(1)

BG    = "#1a1f2e"
DARK  = "#0d1117"
PANEL = "#252d3e"
INPUT = "#0f1520"
GOLD  = "#e8c96d"
BLUE  = "#7eb8f7"
GREEN = "#5fba7d"
MUTED = "#4a5568"
RED   = "#e05c5c"


class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Unicode → KrutiDev 10 Converter")
        self.configure(bg=BG)
        self.geometry("640x480")
        self.minsize(540, 400)
        self.resizable(True, True)

        self._in_path  = tk.StringVar()
        self._out_path = tk.StringVar()
        self._status   = tk.StringVar(value="Select an Excel file to begin.")
        self._progress = tk.IntVar(value=0)

        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Title bar ─────────────────────────────────────────────────
        bar = tk.Frame(self, bg=DARK, pady=12)
        bar.pack(fill="x")
        tk.Label(bar,
                 text="🔤  Unicode → KrutiDev 10 Converter",
                 font=("Georgia", 13, "bold"), bg=DARK, fg=GOLD,
                 padx=16).pack(side="left")

        # ── Scrollable body ───────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=12)

        # ── Input file ────────────────────────────────────────────────
        tk.Label(body, text="📂  Input Excel file",
                 font=("Helvetica", 10, "bold"), bg=BG, fg=BLUE,
                 anchor="w").pack(fill="x", pady=(0, 4))

        in_row = tk.Frame(body, bg=BG)
        in_row.pack(fill="x", pady=(0, 12))
        tk.Entry(in_row, textvariable=self._in_path,
                 font=("Courier", 9),
                 bg=INPUT, fg="#d4e8ff", insertbackground="#d4e8ff",
                 relief="flat", bd=4
                 ).pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(in_row, text="Browse…",
                  font=("Helvetica", 9), bg=PANEL, fg=BLUE,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._browse_input
                  ).pack(side="left", padx=(8, 0))

        # ── Output file ───────────────────────────────────────────────
        tk.Label(body, text="💾  Output Excel file",
                 font=("Helvetica", 10, "bold"), bg=BG, fg=BLUE,
                 anchor="w").pack(fill="x", pady=(0, 4))

        out_row = tk.Frame(body, bg=BG)
        out_row.pack(fill="x", pady=(0, 14))
        tk.Entry(out_row, textvariable=self._out_path,
                 font=("Courier", 9),
                 bg=INPUT, fg="#d4e8ff", insertbackground="#d4e8ff",
                 relief="flat", bd=4
                 ).pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(out_row, text="Browse…",
                  font=("Helvetica", 9), bg=PANEL, fg=BLUE,
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._browse_output
                  ).pack(side="left", padx=(8, 0))

        # ── Info box ──────────────────────────────────────────────────
        info = tk.Frame(body, bg=PANEL, bd=1, relief="groove")
        info.pack(fill="x", pady=(0, 16))
        tk.Label(info,
                 text="All columns containing Hindi (Devanagari) text are detected "
                      "automatically.  A new _KD column is inserted beside each "
                      "source column and formatted with Kruti Dev 010 font.",
                 font=("Helvetica", 9), bg=PANEL, fg=MUTED,
                 justify="left", padx=12, pady=10,
                 wraplength=520, anchor="w").pack(fill="x")

        # ── Buttons ───────────────────────────────────────────────────
        btn_row = tk.Frame(body, bg=BG)
        btn_row.pack(fill="x", pady=(0, 14))
        tk.Button(btn_row,
                  text="🔄  Convert",
                  font=("Helvetica", 11, "bold"),
                  bg="#2ecc71", fg="#1a1a2e",
                  activebackground="#27ae60",
                  relief="flat", cursor="hand2",
                  padx=28, pady=10,
                  command=self._start_convert
                  ).pack(side="left")
        tk.Button(btn_row,
                  text="💾  Save As…",
                  font=("Helvetica", 9),
                  bg="#9b59b6", fg="white",
                  activebackground="#8e44ad",
                  relief="flat", cursor="hand2",
                  padx=14, pady=10,
                  command=self._save_as
                  ).pack(side="left", padx=(12, 0))
        tk.Button(btn_row,
                  text="📂  Open output folder",
                  font=("Helvetica", 9),
                  bg=PANEL, fg=BLUE,
                  relief="flat", cursor="hand2",
                  padx=14, pady=10,
                  command=self._open_folder
                  ).pack(side="left", padx=(12, 0))

        # ── Progress bar ──────────────────────────────────────────────
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("Green.Horizontal.TProgressbar",
                        troughcolor=DARK, background=GREEN, thickness=16)
        self._pb = ttk.Progressbar(body,
                                   variable=self._progress,
                                   maximum=100,
                                   style="Green.Horizontal.TProgressbar")
        self._pb.pack(fill="x", pady=(0, 8))

        # ── Status label ──────────────────────────────────────────────
        tk.Label(body, textvariable=self._status,
                 font=("Courier", 9), bg=BG, fg=MUTED,
                 anchor="w", wraplength=560,
                 justify="left").pack(fill="x")

    def _file_section(self, parent, label, var, browse_cmd, row):
        # Kept for compatibility — not used in new layout
        pass

    # ── File picking ──────────────────────────────────────────────────────────

    def _browse_input(self):
        p = askopenfilename(
            title="Select input Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")])
        if not p:
            return
        self._in_path.set(p)
        # Auto-suggest output path
        stem = Path(p).stem
        out  = str(Path(p).parent / f"{stem}_KrutiDev.xlsx")
        self._out_path.set(out)
        self._status.set(f"Input: {Path(p).name}  —  click Convert to begin.")

    def _browse_output(self):
        p = asksaveasfilename(
            title="Save output as…",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=Path(self._in_path.get()).stem + "_KrutiDev.xlsx"
                if self._in_path.get() else "output_KrutiDev.xlsx")
        if p:
            self._out_path.set(p)

    def _save_as(self):
        """Choose a new save path then run the conversion directly to that path."""
        inp = self._in_path.get().strip().strip('"')
        if not inp:
            messagebox.showwarning("No input", "Please select an input Excel file first.")
            return
        p = asksaveasfilename(
            title="Save converted file as…",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=Path(inp).stem + "_KrutiDev.xlsx")
        if p:
            self._out_path.set(p)
            self._start_convert()

    # ── Conversion ────────────────────────────────────────────────────────────

    def _start_convert(self):
        inp = self._in_path.get().strip().strip('"')
        out = self._out_path.get().strip().strip('"')

        if not inp:
            messagebox.showwarning("No input", "Please select an input Excel file.")
            return
        if not os.path.isfile(inp):
            messagebox.showerror("Not found", f"Cannot find:\n{inp}")
            return
        if not out:
            messagebox.showwarning("No output", "Please choose an output file path.")
            return

        self._progress.set(0)
        self._status.set("⏳  Starting conversion…")
        self.update_idletasks()

        # Run in a thread so the UI stays responsive
        t = threading.Thread(target=self._run_convert,
                             args=(inp, out), daemon=True)
        t.start()

    def _run_convert(self, inp, out):
        def _prog(pct, msg):
            self._progress.set(pct)
            self._status.set(msg)
            self.update_idletasks()

        try:
            rows, cols = process_excel(inp, out, progress_cb=_prog)
            self._status.set(
                f"✅  Done!  {rows:,} rows · {cols} column(s) converted  →  {Path(out).name}")
            self._progress.set(100)
            # Ask to open the file
            self.after(200, lambda: self._offer_open(out))
        except Exception as e:
            self._status.set(f"❌  Error: {e}")
            self._progress.set(0)
            self.after(0, lambda: messagebox.showerror("Conversion failed", str(e)))

    def _offer_open(self, path):
        if messagebox.askyesno("Done", f"Conversion complete.\n\nOpen {Path(path).name}?"):
            self._open_file(path)

    def _open_file(self, path):
        import subprocess
        try:
            if sys.platform == "win32":
                os.startfile(path)              # type: ignore
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception:
            pass

    def _open_folder(self):
        out = self._out_path.get().strip()
        if out:
            folder = str(Path(out).parent)
        else:
            folder = str(Path.home())
        self._open_file(folder)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()